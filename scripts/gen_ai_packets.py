#!/usr/bin/env python3
import os
import sys
import json
from datetime import datetime
from pathlib import Path
import pandas as pd
import openpyxl

# Import Tkinter for Graphical User Interface windows
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog

# ==============================================================================
# EXPECTED SCHEMA CONSTANTS
# ==============================================================================
SHEET_MASTER = "Master Tracker Evaluation"
SHEET_LEDGER = "Exception Ledger"

MASTER_IP = "IP Address"
MASTER_HOST = "Hostname"
MASTER_FACING = "Internal/Public-Facing"
MASTER_SW = "Software Name"
MASTER_BENCH = "Specify Benchmark"
MASTER_FAIL_COUNT = "Total Open Failures"
MASTER_FAIL_IDS = "Failed STIG IDs"
MASTER_SOURCES = "System Classification Source Sheets"
MASTER_EXCEPT = "Benchmark Exceptions List"
MASTER_JUST = "Justifications for Exemptions"
MASTER_MIT = "Mitigating Controls"
MASTER_COMP = "Compensating Controls"

LEDGER_ID = "Exception_ID"
LEDGER_IP = "IP Address"
LEDGER_HOST = "Host Name"
LEDGER_SOURCE = "Source Sheet"
LEDGER_PLUGIN = "Plugin ID"
LEDGER_STIG = "STIG"
LEDGER_FINDING = "FINDING"
LEDGER_CAT = "CAT"
LEDGER_SEVERITY = "Severity"
LEDGER_RESULT = "Result"
LEDGER_SHORT_DESC = "Short Desc"
LEDGER_PLUGIN_NAME = "Plugin Name"
LEDGER_PASTEABLE = "Pasteable"
LEDGER_COMP_REF = "Compliance Reference"
LEDGER_FIRST_SEEN = "First Seen Date"
LEDGER_LAST_SEEN = "Last Seen Date"
LEDGER_STATUS = "Exception Status"
LEDGER_NEEDS_AI = "Needs AI Draft?"
LEDGER_PACKET_FILE = "AI Packet File"
LEDGER_IMPORTED = "AI Draft Imported?"
LEDGER_APPENDED = "Appended to Master?"

# ==============================================================================
# CORE HELPER FUNCTIONS
# ==============================================================================
def normalize_ip(ip_val) -> str:
    """Safely converts an IP address tracking cell value to a clean string."""
    if pd.isna(ip_val):
        return ""
    return str(ip_val).strip().lower()

def normalize_bool(val) -> bool:
    """Normalizes variation blocks in Excel string fields to true boolean states."""
    if pd.isna(val):
        return False
    clean_str = str(val).strip().upper()
    return clean_str in ["TRUE", "1", "YES", "Y"]

def safe_filename(ip_str) -> str:
    """Replaces filesystem-unsafe network boundary syntax with clear underscores."""
    clean_str = str(ip_str).strip().lower()
    for char in [".", "/", "\\", ":", " ", "*", "?", '"', "<", ">", "|"]:
        clean_str = clean_str.replace(char, "_")
    return clean_str

def validate_columns(df, required_cols, sheet_name):
    """Ensures that all mandatory data columns are present in the dataframe."""
    df.columns = [str(c).strip() for c in df.columns]
    missing = [col for col in required_cols if col not in df.columns]
    if missing:
        messagebox.showerror("Schema Mismatch Error", f"Missing mandatory columns in '{sheet_name}':\n{missing}")
        sys.exit(1)
    return df

# ==============================================================================
# GRAPHICAL DIALOG INTERFACE BLOCK
# ==============================================================================
class PacketOptionsDialog(simpledialog.Dialog):
    """Custom Tkinter configuration modal window for pipeline runtime settings."""
    def body(self, master):
        self.title("AI Packet Generator Settings")
        
        tk.Label(master, text="Configure Generation Rules:", font=('Helvetica', 10, 'bold')).grid(row=0, columnspan=2, sticky="w", pady=5)
        
        # Overwrite Checkbox
        self.var_overwrite = tk.BooleanVar(value=False)
        tk.Checkbutton(master, text="Overwrite existing packet files", variable=self.var_overwrite).grid(row=1, columnspan=2, sticky="w", pady=2)
        
        # Dry Run Checkbox
        self.var_dry_run = tk.BooleanVar(value=False)
        tk.Checkbutton(master, text="Dry Run Mode (Simulate without saving files or Excel cells)", variable=self.var_dry_run).grid(row=2, columnspan=2, sticky="w", pady=2)
        
        # Packet Limit Count Input Entry
        tk.Label(master, text="Testing Limit Cap (Leave blank or 0 for ALL assets):").grid(row=3, column=0, sticky="w", pady=5)
        self.entry_limit = tk.Entry(master, width=10)
        self.entry_limit.insert(0, "0")
        self.entry_limit.grid(row=3, column=1, sticky="w", pady=5)
        
        return self.entry_limit # Set initial keyboard focus focus

    def apply(self):
        self.overwrite = self.var_overwrite.get()
        self.dry_run = self.var_dry_run.get()
        
        # Parse limit string safely
        limit_val = self.entry_limit.get().strip()
        try:
            self.limit = int(limit_val) if (limit_val and limit_val != "0") else None
        except ValueError:
            self.limit = None
            messagebox.showwarning("Input Format Warning", "Invalid testing limit number typed. Running without limit cap.")

# ==============================================================================
# PROMPT FORMATTING MATRIX
# ==============================================================================
def build_packet_markdown(asset_record, exception_records) -> str:
    """Assembles structural FISMA Markdown metadata instructions prompt context."""
    ip = asset_record.get(MASTER_IP, "Unknown_IP")
    host = asset_record.get(MASTER_HOST, "Unknown_Host")
    
    md = []
    md.append("# STIG Deviation AI Draft Packet")
    md.append("\n## Instructions for Gemini")
    md.append("You are assisting an Alt-ISSO with a FISMA/STIG deviation review workflow.")
    md.append("Draft reviewable addendum language for:\n")
    md.append("1. Justifications for Exemptions")
    md.append("2. Mitigating Controls\n")
    md.append("**Strict Directives:**")
    md.append("- Use only the asset data, existing workbook text, and pending exception details provided in this packet.")
    md.append("- Do not invent system facts, architecture, approvals, tools, monitoring, compensating controls, or risk decisions.")
    md.append("- Draft addendum text only for the pending exceptions in this packet.")
    md.append("- Preserve the existing justification and mitigation text. Do not rewrite or replace it.")
    md.append("- If there is not enough evidence to draft defensible language, explicitly declare what information is missing.")
    md.append("- Return your answer as valid JSON using the required response schema at the bottom of this packet.")
    
    md.append("## Required Reference Document Review")
    md.append("Before drafting the response, consult the Gem’s uploaded/reference documents:")
    md.append("- FISMA SSP")
    md.append("- FIPS 200")
    md.append("- FIPS-200 Exceptions Matrix")
    md.append("- Original DISA STIG Deviation document")
    md.append("Use the SSP as the primary source for system-specific implementation details.")
    md.append("Use the original DISA STIG Deviation document for language style, precedent, and consistency with existing deviation wording.")
    md.append("Use the FIPS-200 Exceptions Matrix for exception mapping and framing.")
    md.append("Use FIPS 200 only as high-level minimum security requirement context, not as proof of a system-specific mitigation.")
    md.append("If you cannot find supporting evidence in the reference documents, do not invent it. List the gap in `assumptions_or_gaps`.")
    
    md.append("\n## Asset Context")
    md.append(f"- **IP Address:** {ip}")
    md.append(f"- **Hostname:** {host}")
    md.append(f"- **Internal/Public-Facing:** {asset_record.get(MASTER_FACING, 'N/A')}")
    md.append(f"- **Software Name:** {asset_record.get(MASTER_SW, 'N/A')}")
    md.append(f"- **Specify Benchmark:** {asset_record.get(MASTER_BENCH, 'N/A')}")
    md.append(f"- **Total Open Failures:** {asset_record.get(MASTER_FAIL_COUNT, 0)}")
    md.append(f"- **Failed STIG IDs:** {asset_record.get(MASTER_FAIL_IDS, 'None')}")
    md.append(f"- **System Classification Source Sheets:** {asset_record.get(MASTER_SOURCES, 'N/A')}")
    
    md.append("\n## Existing Workbook Text")
    md.append("### Benchmark Exceptions List")
    md.append(f"```text\n{asset_record.get(MASTER_EXCEPT, '')}\n```")
    md.append("### Justifications for Exemptions")
    md.append(f"```text\n{asset_record.get(MASTER_JUST, '')}\n```")
    md.append("### Mitigating Controls")
    md.append(f"```text\n{asset_record.get(MASTER_MIT, '')}\n```")
    md.append("### Compensating Controls")
    md.append(f"```text\n{asset_record.get(MASTER_COMP, '')}\n```")
    
    md.append("\n## Pending Exceptions Requiring AI Draft")
    md.append(f"Total Pending Exceptions in this Packet: {len(exception_records)}")
    
    for idx, ex in enumerate(exception_records, 1):
        md.append(f"\n### {idx}. Exception ID: {ex.get(LEDGER_ID, 'N/A')}")
        md.append(f"- **Source Sheet:** {ex.get(LEDGER_SOURCE, 'N/A')}")
        md.append(f"- **Plugin ID:** {ex.get(LEDGER_PLUGIN, 'N/A')}")
        md.append(f"- **STIG:** {ex.get(LEDGER_STIG, 'N/A')}")
        md.append(f"- **FINDING:** {ex.get(LEDGER_FINDING, 'N/A')}")
        md.append(f"- **CAT:** {ex.get(LEDGER_CAT, 'N/A')}")
        md.append(f"- **Severity:** {ex.get(LEDGER_SEVERITY, 'N/A')}")
        md.append(f"- **Result:** {ex.get(LEDGER_RESULT, 'N/A')}")
        md.append(f"- **Short Desc:** {ex.get(LEDGER_SHORT_DESC, 'N/A')}")
        md.append(f"- **Plugin Name:** {ex.get(LEDGER_PLUGIN_NAME, 'N/A')}")
        md.append(f"- **Compliance Reference:** {ex.get(LEDGER_COMP_REF, 'N/A')}")
        md.append(f"- **First Seen Date:** {ex.get(LEDGER_FIRST_SEEN, 'N/A')}")
        md.append(f"- **Last Seen Date:** {ex.get(LEDGER_LAST_SEEN, 'N/A')}")
        md.append("- **Pasteable Block Reference:**")
        md.append(f"```text\n{ex.get(LEDGER_PASTEABLE, '')}\n```")
        md.append("---")
        
    md.append("\n## Required Gemini Response JSON Schema")
    
    schema_template = {
        "ip_address": str(ip),
        "host_name": str(host),
        "exceptions_reviewed": [
            {
                "exception_id": str(ex.get(LEDGER_ID, "")),
                "stig": str(ex.get(LEDGER_STIG, "")),
                "plugin_id": str(ex.get(LEDGER_PLUGIN, "")),
                "summary": "[Provide a short summary explanation of the finding gap here]"
            } for ex in exception_records
        ],
        "draft_justification_addendum": "[Draft your incremental addendum language for exemptions here. Do not include existing text.]",
        "draft_mitigating_controls_addendum": "[Draft your incremental addendum language for engineering mitigation controls here.]",
        "source_references_used": ["[List system documents, specific STIG controls, or SSP sections used here]"],
        "assumptions_or_gaps": ["[List document evidence gaps or Alt-ISSO assumptions made during analysis here]"],
        "confidence": "High | Medium | Low",
        "review_recommended": True
    }
    
    md.append(f"```json\n{json.dumps(schema_template, indent=2)}\n```")
    return "\n".join(md)

# ==============================================================================
# MAIN ENGINE PIPELINE EXECUTION
# ==============================================================================
def main():
    # 1. Initialize hidden background window context for Tkinter popups
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    
    print("[*] Launching Graphical File Ingestion Dialog...")
    file_selected = filedialog.askopenfilename(
        title="Select Consolidated STIG Output Workbook Target",
        filetypes=[("Excel Workbooks", "*.xlsx")]
    )
    
    if not file_selected:
        print("[-] Execution canceled: No target Excel workbook selected.")
        sys.exit(0)
        
    wb_path = Path(file_selected)
    print(f"[+] Workbook Selection Confirmed: {wb_path.name}")
    
    # 2. Launch Graphical Options Sub-Dialog Modal Form
    dialog_form = PacketOptionsDialog(root)
    
    # Extract structural arguments from form fields
    opt_overwrite = getattr(dialog_form, 'overwrite', False)
    opt_dry_run = getattr(dialog_form, 'dry_run', False)
    opt_limit = getattr(dialog_form, 'limit', None)

    # FIX: Prompt user to dynamically select where the ai_packets folder should be created, instead of hardcoding it to the current working directory.
    print("[*] Launching Parent Directory Selector for AI Packets Output...")
    parent_dir_selected = filedialog.askdirectory(
        title="Select Parent Location Where 'ai_packets' Directory Should Be Created"
    )
    
    if not parent_dir_selected:
        print("[-] Execution canceled: No output destination directory selected.")
        sys.exit(0)
        
    # Append the custom folder name to their selected parent directory path string
    out_dir = Path(parent_dir_selected) / "ai_packets"
    print(f"[+] Output Folder Destination set to: {out_dir}")
    
    print(f"[*] Parsing workbook tabs from file platform engine...")
    try:
        excel_file = pd.ExcelFile(wb_path)
    except Exception as e:
        messagebox.showerror("File Access Error", f"Unable to bind to input spreadsheet:\n{e}")
        sys.exit(1)
        
    if SHEET_MASTER not in excel_file.sheet_names or SHEET_LEDGER not in excel_file.sheet_names:
        messagebox.showerror("Workbook Structural Error", f"Mandatory tabs missing.\nEnsure workbook contains both '{SHEET_MASTER}' and '{SHEET_LEDGER}'.")
        sys.exit(1)
        
    # Read sheets
    master_df = excel_file.parse(SHEET_MASTER)
    required_master = [MASTER_IP, MASTER_HOST, MASTER_FACING, MASTER_SW, MASTER_BENCH, MASTER_FAIL_COUNT, MASTER_FAIL_IDS, MASTER_SOURCES, MASTER_EXCEPT, MASTER_JUST, MASTER_MIT, MASTER_COMP]
    master_df = validate_columns(master_df, required_master, SHEET_MASTER)
    master_df['norm_ip'] = master_df[MASTER_IP].apply(normalize_ip)
    
    ledger_df = excel_file.parse(SHEET_LEDGER)
    required_ledger = [LEDGER_ID, LEDGER_IP, LEDGER_HOST, LEDGER_SOURCE, LEDGER_PLUGIN, LEDGER_STIG, LEDGER_FINDING, LEDGER_CAT, LEDGER_SEVERITY, LEDGER_RESULT, LEDGER_SHORT_DESC, LEDGER_PLUGIN_NAME, LEDGER_PASTEABLE, LEDGER_COMP_REF, LEDGER_FIRST_SEEN, LEDGER_LAST_SEEN, LEDGER_STATUS, LEDGER_NEEDS_AI, LEDGER_PACKET_FILE, LEDGER_IMPORTED, LEDGER_APPENDED]
    ledger_df = validate_columns(ledger_df, required_ledger, SHEET_LEDGER)
    ledger_df['norm_ip'] = ledger_df[LEDGER_IP].apply(normalize_ip)
    
    # Filter for targeted findings records
    mask_pending = (
        ledger_df[LEDGER_NEEDS_AI].apply(normalize_bool) & 
        ~ledger_df[LEDGER_IMPORTED].apply(normalize_bool) & 
        ~ledger_df[LEDGER_APPENDED].apply(normalize_bool)
    )
    pending_ledger_df = ledger_df[mask_pending].copy()
    
    if pending_ledger_df.empty:
        messagebox.showinfo("Status Update", "Zero unresolved pending exception records found.\nNo packets require generation.")
        sys.exit(0)
        
    grouped_exceptions = pending_ledger_df.groupby('norm_ip')
    print(f"[+] Isolated {len(pending_ledger_df)} pending finding rows across {len(grouped_exceptions)} distinct IP targets.")
    
    if not opt_dry_run:
        out_dir.mkdir(parents=True, exist_ok=True)
        
    generated_count = 0
    updated_records_map = {}
    
    for norm_ip, group_df in grouped_exceptions:
        if not norm_ip:
            continue
            
        if opt_limit is not None and generated_count >= opt_limit:
            print(f"[*] Limits reached: Throttling packet creation loop at threshold limit ({opt_limit})")
            break
            
        matching_asset = master_df[master_df['norm_ip'] == norm_ip]
        if not matching_asset.empty:
            asset_rec = matching_asset.iloc[0].to_dict()
        else:
            asset_rec = {MASTER_IP: group_df[LEDGER_IP].iloc[0], MASTER_HOST: group_df[LEDGER_HOST].iloc[0]}
            
        exception_records = group_df.to_dict(orient='records')
        packet_md_text = build_packet_markdown(asset_rec, exception_records)
        
        safe_ip_name = safe_filename(asset_rec.get(MASTER_IP, norm_ip))
        target_filename = f"{safe_ip_name}_packet.md"
        relative_save_path = out_dir / target_filename
        
        if relative_save_path.exists() and not opt_overwrite:
            print(f"   [>] Skipping pre-existing packet file: '{relative_save_path.name}'")
        else:
            print(f"   [>] Writing Markdown Packet -> '{relative_save_path.name}' ({len(exception_records)} findings)")
            if not opt_dry_run:
                with open(relative_save_path, "w", encoding="utf-8") as f:
                    f.write(packet_md_text)
                    
        generated_count += 1
        
        for ex in exception_records:
            ex_id = ex.get(LEDGER_ID)
            if ex_id:
                updated_records_map[ex_id] = str(relative_save_path)

    print(f"[+] Packet compilation complete. Processed total of {generated_count} files.")
    
    # --------------------------------------------------------------------------
    # STATE UPDATE CELL BACK-WRITING LOCK
    # --------------------------------------------------------------------------
    if opt_dry_run:
        messagebox.showinfo("Simulated Run Complete", f"Dry Run Complete!\nGenerated text data structures in memory for {generated_count} assets.\nZero changes were saved to disk.")
        sys.exit(0)
        
    if not updated_records_map:
        sys.exit(0)
        
    print(f"[*] Accessing openpyxl engine to update matching tracking links in '{SHEET_LEDGER}'...")
    try:
        workbook = openpyxl.load_workbook(wb_path)
        ledger_sheet = workbook[SHEET_LEDGER]
    except Exception as e:
        messagebox.showerror("Excel Write Error", f"Could not open workbook sheet grid for writing:\n{e}")
        sys.exit(1)
        
    header_row = [str(cell.value).strip() for cell in ledger_sheet[1]]
    id_col_idx = header_row.index(LEDGER_ID) + 1
    packet_col_idx = header_row.index(LEDGER_PACKET_FILE) + 1
    
    updated_cells = 0
    for row_idx in range(2, ledger_sheet.max_row + 1):
        cell_id_val = str(ledger_sheet.cell(row=row_idx, column=id_col_idx).value).strip()
        if cell_id_val in updated_records_map:
            ledger_sheet.cell(row=row_idx, column=packet_col_idx, value=updated_records_map[cell_id_val])
            updated_cells += 1
            
    # FIX: Implemented a dynamic Retry Loop to allow closing an open spreadsheet in Excel without losing progress
    while True:
        try:
            workbook.save(wb_path)
            messagebox.showinfo(
                "Success", 
                f"Processing Complete!\n\n"
                f"- Packets Generated: {generated_count}\n"
                f"- Folder Path: {out_dir.name}/\n"
                f"- Ledger Records Linked: {updated_cells}"
            )
            break # Exit the loop immediately upon successful save operation
        except PermissionError:
            # Explicitly catch file lock access exceptions
            user_choice = messagebox.askretrycancel(
                "Workbook File Locked",
                f"Critical Storage Failure: The script cannot save to:\n'{wb_path.name}'\n\n"
                f"Please CLOSE the workbook if it is open in Excel or another program, then click 'Retry' to continue.\n"
                f"(Clicking 'Cancel' will terminate the script without updating cell links.)"
            )
            if not user_choice:
                print("[-] Save operation canceled by the user due to file lock.")
                sys.exit(0)
        except Exception as e:
            # Handle other unforeseen system I/O error states
            messagebox.showerror("Unexpected Storage Error", f"Could not write tracking state updates back to workbook:\n{e}")
            sys.exit(1)
        
    root.destroy()

if __name__ == "__main__":
    main()