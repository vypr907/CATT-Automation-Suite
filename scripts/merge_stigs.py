import os
import sys
import json
from datetime import datetime
from pathlib import Path
import pandas as pd
import hashlib

# Import Tkinter for Graphical File Dialog Windows
import tkinter as tk
from tkinter import filedialog, messagebox

# Import Openpyxl styling utilities for formatting the output workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Try loading optional integrations for Google Sheets
try:
    import gspread
    from oauth2client.service_account import Credentials
    GSPREAD_AVAILABLE = True
except ImportError:
    GSPREAD_AVAILABLE = False

# ==============================================================================
# CONFIGURATION & COLUMN MAPPING
# ==============================================================================
# Deviation Tracker Column Constants (Master Sheet)
DEV_IP_COL = "IP Address"
DEV_HOST_COL = "Host Name"
DEV_FACING_COL = "Host is Internal or Public-Facing" # Matches exact spacing
DEV_TYPE_COL = "Host Type"
DEV_SW_COL = "Software Name"
DEV_BENCH_COL = "Specify Benchmark followed"
DEV_EXCEPT_COL = "Benchmark Exceptions List"
DEV_JUST_COL = "Justifications for Exemptions"
DEV_MIT_COL = "Mitigating Controls"
DEV_COMP_COL = "Compensating Controls"

# Incoming Scan Column Constants (Raw Scan Data)
SCAN_HOST_COL = "hostname"  # This column holds IP addresses strings in incoming data
SCAN_STIG_COL = "stig"
SCAN_RESULT_COL = "result"
SCAN_PLUGIN_COL = "plugin id"
SCAN_PLUGIN_NAME = "plugin name"
SCAN_SHORT_DESC = "short description"
SCAN_PASTEABLE = "pasteable"

def normalize_ip(ip_val) -> str:
    """Helper function to normalize IP addresses/join keys uniformly"""
    if pd.isna(ip_val):
        return ""
    return str(ip_val).strip().lower()

def is_active_failure(result_val) -> bool:
    """Helper function to determine if a scan result indicates an active failure"""
    if pd.isna(result_val):
        return False
    return str(result_val).strip().upper() in ["FAILED", "FAIL"]

def clean_join_list(items_list, separator=", ") -> str:
    """Helper function to clean, deduplicate, and merge readable text arrays / join a list of items into a single string"""
    cleaned = sorted(list(set([str(x).strip() for x in items_list if pd.notna(x) and str(x).strip() != ""])))
    return separator.join(cleaned)

def generate_composite_key(ip_or_host: str, stig_id: str) -> str:
    """Generates a tracking key combining asset and rule signature."""
    clean_ip = str(ip_or_host).strip().lower()
    clean_stig = str(stig_id).strip().split()[0].split(',')[0].upper()
    return f"{clean_ip}::{clean_stig}"

def find_hostname_column(columns) -> str:
    """Dynamically locates the asset/hostname identifier column matching variants."""
    normalized_cols = {str(c).strip().lower().replace(" ", "").replace("_", ""): str(c) for c in columns}
    # Look for cleaned common variations found in Nessus/Tenable and your tracker
    for variant in ["hostname", "ipaddress", "hostname", "host", "ip"]:
        if variant in normalized_cols:
            return normalized_cols[variant]
    return ""

def generate_exception_id(ip, plugin_id, stig, finding, pasteable) -> str:
    """Generates a stable, deterministic MD5 hash string for an exception signature."""
    signature = f"{str(ip).strip().lower()}||{str(plugin_id).strip().lower()}||{str(stig).strip().lower()}||{str(finding).strip().lower()}||{str(pasteable).strip().lower()}"
    return hashlib.md5(signature.encode('utf-8')).hexdigest()

def is_already_in_exceptions(pasteable_val, base_exceptions_text) -> bool:
    """Checks if a pasteable entry text block already exists in the exceptions string case-insensitively."""
    if pd.isna(base_exceptions_text) or str(base_exceptions_text).strip() == "":
        return False
    if pd.isna(pasteable_val) or str(pasteable_val).strip() == "":
        return False
    return str(pasteable_val).strip().lower() in str(base_exceptions_text).strip().lower()

def load_from_google_sheets(target_input: str) -> dict[str, pd.DataFrame]:
    """Authenticates and pulls sheets from Google Drive environment, returning a mapping of sheet names to dataframes."""
    if not GSPREAD_AVAILABLE:
        print("[-] Error: 'gspread' or 'google-auth' dependencies missing. Run: pip install gspread google-auth")
        sys.exit(1)

    creds_path = Path("service_account.json")
    if not creds_path.exists():
        print(f"[-] Authentication Failure: Cloud credentials missing. Place '{creds_path}' in this directory.")
        sys.exit(1)

    print("[*] Connecting to Google Sheets API using service account credentials...")
    scopes = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
    creds = Credentials.from_service_account_file(str(creds_path), scopes=scopes)
    gc = gspread.authorize(creds)
    
    try:
        # Check if URL or exact sheet title/ID was supplied
        if "docs.google.com" in target_input:
            sh = gc.open_by_url(target_input)
        else:
            sh = gc.open(target_input)
            
        sheets_dict = {}
        for worksheet in sh.worksheets():
            records = worksheet.get_all_records()
            if records:
                sheets_dict[worksheet.title] = pd.DataFrame(records)
        return sheets_dict
    except Exception as e:
        print(f"[-] Failed to access Google Sheet '{target_input}': {e}")
        sys.exit(1)

def load_local_file(file_path: Path) -> dict[str, pd.DataFrame]:
    """Loads a local .xlsx or .csv file safely, mapping sheet names to dataframes."""
    print(f"[+] Parsing local File: {file_path.name}")
    suffix = file_path.suffix.lower()

    if suffix == ".xlsx":
        sheets = pd.read_excel(file_path, sheet_name=None)
        # Normalize column names for each sheet to ensure consistent lookups
        for sheet_name in sheets:
            sheets[sheet_name].columns = [str(c).strip() for c in sheets[sheet_name].columns]
        return sheets
    elif suffix == ".csv":
        df = pd.read_csv(file_path)
        # Normalize column names for the CSV to ensure consistent lookups
        df.columns = [str(c).strip() for c in df.columns]
        return {file_path.stem: df}
    else:
        print(f"[-] Unsupported file type '{suffix}' for file: {file_path.name}. Only .xlsx and .csv are supported.")
        sys.exit(1)

def gui_select_source(label: str) -> dict[str, pd.DataFrame]:
    """
    Uses Tkinter to pop up a native system dialog picker box. 
    Allows pasting a Google Sheets URL if the terminal is preferred.
    """
    # Initialize hidden tkinter background window environment
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)  # Forces pop-up folder dialog over terminal
    
    print(f"\n[*] Opening File Explorer Window: Choose [{label}]")
    
    # Prompt user option via popup box
    use_cloud = messagebox.askyesno(
        "Source Ingestion Interface", 
        f"Is the [{label}] hosted on Google Sheets?\n\n(Click 'No' to pick a local .xlsx or .csv file instead)"
    )
    
    if use_cloud:
        # Prompt user to paste link into terminal
        print(f"    --> Please paste the Google Sheets URL or Exact Sheet Title into the terminal below:")
        cloud_input = input(">> Paste Google Sheet URL/Name: ").strip()
        if not cloud_input:
            print("[-] Error: URL or name cannot be empty.")
            sys.exit(1)
        root.destroy()
        return load_from_google_sheets(cloud_input)
    else:
        # Launch Graphical Local File Picker window
        file_selected = filedialog.askopenfilename(
            title=f"Select {label} File",
            filetypes=[("Excel or CSV Files", "*.xlsx;*.csv"), ("Excel Workbooks", "*.xlsx"), ("Comma Separated Values", "*.csv")]
        )
        
        if not file_selected:
            print(f"[-] File selection canceled for [{label}]. Exiting pipeline.")
            root.destroy()
            sys.exit(0)
            
        print(f"[+] Graphical Selection Confirmed: {Path(file_selected).name}")
        root.destroy()
        return load_local_file(Path(file_selected))

def gui_select_save_destination(default_filename: str) -> Path:
    """Pops up a Graphical Save Dialog Window for selecting output path destination."""
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    
    print("\n[*] Opening File Explorer Window: Choose destination folder to SAVE report...")
    file_destination = filedialog.asksaveasfilename(
        title="Save Consolidated STIG Report",
        initialfile=default_filename,
        defaultextension=".xlsx",
        filetypes=[("Excel Workbook", "*.xlsx")]
    )
    
    if not file_destination:
        print("[-] Save operation canceled by operator. Defaulting to local working directory instead.")
        root.destroy()
        return Path.cwd() / default_filename
        
    root.destroy()
    return Path(file_destination)

def merge_deviation_sheets():
    print(f"[*] Starting STIG pipeline processing algorithm at {datetime.now()}")
    
    # Intialize Tkinter hidden Root Window for GUI dialogs
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)  # Forces pop-up folder dialog over terminal

    #-----------------------------------------------
    # STEP 1 & 2: Ingest and Normalize Data Sources
    #-----------------------------------------------


    print(f"\n[*] Opening File Explorer Window: Choose [Master Deviation Tracker Spreadsheet]")
    master_file = filedialog.askopenfilename(
        title="Select Master Deviation Tracker Spreadsheet",
        filetypes=[("Excel Workbooks", "*.xlsx")]
    )
    if not master_file:
        print("[-] Execution canceled: Master tracker file not provided.")
        sys.exit(0)
        
    master_df = pd.read_excel(master_file, sheet_name=0)
    # Fix: Clean whitespaces from column names to guarantee lookups work
    master_df.columns = [str(c).strip() for c in master_df.columns]
    print(f"[+] Loaded Master Tracker Sheet: {len(master_df)} rows found.")
    
    print(f"\n[*] Opening File Explorer Window: Choose [Incoming Raw Scan Workbooks Pool]")
    scan_file = filedialog.askopenfilename(
        title="Select Incoming Raw Scan Workbook Pool",
        filetypes=[("Excel Workbooks", "*.xlsx")]
    )
    if not scan_file:
        print("[-] Execution canceled: Scan data file not provided.")
        sys.exit(0)
        
    scan_excel = pd.ExcelFile(scan_file)
    scan_sheets_dict = {sheet: scan_excel.parse(sheet) for sheet in scan_excel.sheet_names}
    
    # --------------------------------------------------------------------------
    # SHEET 2 Generation: Raw Scan Data Pool (Grain: One row per scan finding)
    # --------------------------------------------------------------------------
    combined_scan_list = []

    # Pool tracking records across sheets dynamically
    for sheet_name, df in scan_sheets_dict.items():
        # FIX: Check against a lowercased list of columns to ensure the expected hostname column exists
        df_cols_lower = [str(c).strip().lower() for c in df.columns]
        if df.empty or SCAN_HOST_COL not in df_cols_lower:
            continue
        df = df.copy()
        # FIX: Clean whitespace AND lowercase scan headers from column names to guarantee lookups work
        df.columns = [str(c).strip().lower() for c in df.columns]
        df['source sheet'] = sheet_name
        combined_scan_list.append(df)
        
    if not combined_scan_list:
        print("[-] Critical Error: No matching scan sheets could be constructed.")
        sys.exit(1)
        
    raw_scan_pool_df = pd.concat(combined_scan_list, ignore_index=True)
    print(f"[+] Combined Authoritative Raw Scan Data Pool (Sheet 2): {len(raw_scan_pool_df)} rows.")

    # Apply global string normalization values to tracking arrays
    raw_scan_pool_df['norm_ip'] = raw_scan_pool_df[SCAN_HOST_COL].apply(normalize_ip)
    master_df['norm_ip'] = master_df[DEV_IP_COL].apply(normalize_ip)

    # DEBUG
    print(f"DEBUG: Sample IPs in Scan Pool: {list(raw_scan_pool_df['norm_ip'].unique()[:5])}")
    print(f"DEBUG: Sample IPs in Master Tracker: {list(master_df['norm_ip'].unique()[:5])}")

    # Remove completely empty rows from the Master data frame join array
    master_df = master_df[master_df['norm_ip'] != ""].copy()

    # Get a list of all unique IP addresses across both files
    all_unique_ips = sorted(list(set(master_df['norm_ip'].unique()).union(set(raw_scan_pool_df['norm_ip'].unique()))))

    # --------------------------------------------------------------------------
    # SHEET 1 Generation: Master Tracker Evaluation and Asset Summary
    # Grain: One row per unique IP address / asset
    # --------------------------------------------------------------------------
    print("[*] Compiling Sheet 1 (Master Tracker Evaluation and Asset Summary)...")
    sheet1_rows = []
    
    for ip in all_unique_ips:
        # Filter raw pool data to pull details for this asset
        # FIX: Ensure absolute clean string matching by stripping the target key inline
        clean_ip_target = str(ip).strip().lower()
        
        # Filter raw pool data strictly using the sanitized target string
        ip_scan_findings = raw_scan_pool_df[raw_scan_pool_df['norm_ip'].str.strip() == clean_ip_target]
        ip_failures = ip_scan_findings[ip_scan_findings[SCAN_RESULT_COL].apply(is_active_failure)]
        
        # Calculate asset rollup compliance properties
        total_open_failures = len(ip_failures)

        # FIX: Ensure it leaves the field blank if there are absolutely no scan findings for this asset
        if ip_scan_findings.empty:
            failed_stig_ids = ""
        elif SCAN_STIG_COL in ip_failures.columns:
            failed_stig_ids = clean_join_list(ip_failures[SCAN_STIG_COL].unique(), separator=", ")
        else:
            failed_stig_ids = "Column Missing"
        
        
        source_sheets = clean_join_list(ip_scan_findings['source sheet'].unique(), separator=", ")
        
        # Fetch the matching row from the Master Tracker if it exists
        matching_dev = master_df[master_df['norm_ip'] == ip]
        
        if not matching_dev.empty:
            dev_row = matching_dev.iloc[0]
            # Map manual variables straight from existing records
            orig_ip = dev_row.get(DEV_IP_COL, "")
            orig_host = dev_row.get(DEV_HOST_COL, "")
            orig_facing = dev_row.get(DEV_FACING_COL, "")
            orig_sw = dev_row.get(DEV_SW_COL, "")
            orig_bench = dev_row.get(DEV_BENCH_COL, "")
            base_exceptions = str(dev_row.get(DEV_EXCEPT_COL, "")) if pd.notna(dev_row.get(DEV_EXCEPT_COL, "")) else ""
            orig_just = dev_row.get(DEV_JUST_COL, "")
            orig_mit = dev_row.get(DEV_MIT_COL, "")
            orig_comp = dev_row.get(DEV_COMP_COL, "")
        else:
            # Create default placeholders for newly discovered assets
            orig_ip = ip_scan_findings[SCAN_HOST_COL].iloc[0] if not ip_scan_findings.empty else ip
            orig_host = ""
            orig_facing = ""
            orig_sw = ""
            orig_bench = ""
            base_exceptions = ""
            orig_just = ""
            orig_mit = ""
            orig_comp = ""

        # Smart Append Logic: Extract unique Pasteable text blocks and check if they are already in the base string
        unique_scan_pasteables = sorted(list(ip_scan_findings[SCAN_PASTEABLE].dropna().unique()))
        appended_exceptions = base_exceptions.strip()
        
        for paste_item in unique_scan_pasteables:
            paste_str = str(paste_item).strip()
            if paste_str and paste_str not in appended_exceptions:
                if appended_exceptions:
                    appended_exceptions += "\n" + paste_str
                else:
                    appended_exceptions = paste_str

        # Append row object to Sheet 1 dataset array matching instructions
        sheet1_rows.append({
            "IP Address": orig_ip,
            "Hostname": orig_host,
            "Internal/Public-Facing": orig_facing,
            "Software Name": orig_sw,
            "Specify Benchmark": orig_bench,
            "Total Open Failures": total_open_failures,
            "Failed STIG IDs": failed_stig_ids if failed_stig_ids else "None",
            "System Classification Source Sheets": source_sheets if source_sheets else "Not Seen in Scans",
            "Benchmark Exceptions List": appended_exceptions,
            "Justifications for Exemptions": orig_just,
            "Mitigating Controls": orig_mit,
            "Compensating Controls": orig_comp
        })
        
    sheet1_df = pd.DataFrame(sheet1_rows)

    # --------------------------------------------------------------------------
    # SHEET 3 Generation: STIG ID Summary (Grain: One row per unique STIG ID)
    # --------------------------------------------------------------------------
    print("[*] Compiling Sheet 3 (STIG ID Summary)...")
    sheet3_rows = []
    unique_stigs = [x for x in raw_scan_pool_df[SCAN_STIG_COL].dropna().unique() if str(x).strip() != ""]
    
    for stig_id in sorted(unique_stigs):
        stig_data = raw_scan_pool_df[raw_scan_pool_df[SCAN_STIG_COL] == stig_id]
        stig_failures = stig_data[stig_data[SCAN_RESULT_COL].apply(is_active_failure)]
        
        plugin_ids = clean_join_list(header_variant_map if 'header_variant_map' in locals() else stig_data[SCAN_PLUGIN_COL].unique())
        
        # Pick description column cleanly prioritizing Short Desc, then Plugin Name
        if SCAN_SHORT_DESC in stig_data.columns and stig_data[SCAN_SHORT_DESC].notna().any():
            description = str(stig_data[SCAN_SHORT_DESC].dropna().iloc[0]).strip()
        elif SCAN_PLUGIN_NAME in stig_data.columns and stig_data[SCAN_PLUGIN_NAME].notna().any():
            description = str(stig_data[SCAN_PLUGIN_NAME].dropna().iloc[0]).strip()
        else:
            description = "N/A"
            
        findings_list = clean_join_list(stig_data[SCAN_PASTEABLE].unique(), separator="\n")
        impacted_hosts = clean_join_list(stig_data[SCAN_HOST_COL].unique())
        total_active_failures = len(stig_failures)
        
        sheet3_rows.append({
            "STIG ID": stig_id,
            "Plugin ID(s)": plugin_ids,
            "Rule Title / Description": description,
            "Findings List": findings_list,
            "Impacted Host List": impacted_hosts,
            "Total Active Host Failures": total_active_failures
        })
        
    sheet3_df = pd.DataFrame(sheet3_rows)

    # --------------------------------------------------------------------------
    # SHEET 4 Generation: Plugin ID Summary (Grain: Preserved Plugin ID Logic)
    # --------------------------------------------------------------------------
    print("[*] Compiling Sheet 4 (Plugin ID Summary)...")
    sheet4_rows = []
    unique_plugins = [x for x in raw_scan_pool_df[SCAN_PLUGIN_COL].dropna().unique() if str(x).strip() != ""]
    
    for plugin_id in sorted(unique_plugins):
        plugin_data = raw_scan_pool_df[raw_scan_pool_df[SCAN_PLUGIN_COL] == plugin_id]
        plugin_failures = plugin_data[plugin_data[SCAN_RESULT_COL].apply(is_active_failure)]
        
        associated_stigs = clean_join_list(plugin_data[SCAN_STIG_COL].unique())
        
        if SCAN_PLUGIN_NAME in plugin_data.columns and plugin_data[SCAN_PLUGIN_NAME].notna().any():
            p_desc = str(plugin_data[SCAN_PLUGIN_NAME].dropna().iloc[0]).strip()
        else:
            p_desc = "N/A"
            
        total_evaluated = len(plugin_data)
        total_active_failures = len(plugin_failures)
        impacted_hosts = clean_join_list(plugin_failures[SCAN_HOST_COL].unique())
        
        sheet4_rows.append({
            "Plugin ID": plugin_id,
            "Associated STIG ID(s)": associated_stigs,
            "Rule Title / Description": p_desc,
            "Total Evaluated Items": total_evaluated,
            "Total Active Host Failures": total_active_failures,
            "Impacted Host List": impacted_hosts if impacted_hosts else "None"
        })
        
    sheet4_df = pd.DataFrame(sheet4_rows)

    # --------------------------------------------------------------------------
    # SHEET 5 Generation: Exception Ledger (Grain: Unique asset/finding combination)
    # --------------------------------------------------------------------------
    print("[*] Compiling New Sheet (Exception Ledger)...")
    ledger_dict = {}
    current_run_date = datetime.now().strftime("%Y-%m-%d")
    
    for _, scan_row in raw_scan_pool_df.iterrows():
        # Extrapolate find grain inputs
        s_ip = normalize_ip(scan_row.get(SCAN_HOST_COL, ""))
        s_plugin = str(scan_row.get(SCAN_PLUGIN_COL, "")).strip()
        s_stig = str(scan_row.get(SCAN_STIG_COL, "")).strip()
        s_finding = str(scan_row.get(SCAN_SHORT_DESC, "")).strip() if SCAN_SHORT_DESC in scan_row else "" # Adjusting finding string mapping safely
        # If your data uses another column specifically named "finding", use scan_row.get('finding', '')
        if 'finding' in scan_row:
            s_finding = str(scan_row.get('finding', '')).strip()
            
        s_pasteable = str(scan_row.get(SCAN_PASTEABLE, "")).strip()
        
        if not s_ip:
            continue
            
        # Compute stable tracking key
        exception_id = generate_exception_id(s_ip, s_plugin, s_stig, s_finding, s_pasteable)
        
        # Determine host name lookup from Deviation Master
        matching_master = master_df[master_df['norm_ip'] == s_ip]
        s_hostname = ""
        base_exceptions_blob = ""
        if not matching_master.empty:
            s_hostname = str(matching_master.iloc[0].get(DEV_HOST_COL, "")).strip()
            base_exceptions_blob = str(matching_master.iloc[0].get(DEV_EXCEPT_COL, "")).strip()
            
        source_sheet_name = str(scan_row.get('source sheet', 'Unknown')).strip()
        
        if exception_id in ledger_dict:
            # Combine source sheet lists if finding is found duplicated across inputs
            existing_sheets = [x.strip() for x in ledger_dict[exception_id]['Source Sheet'].split(",")]
            if source_sheet_name not in existing_sheets:
                ledger_dict[exception_id]['Source Sheet'] += f", {source_sheet_name}"
            continue
            
        # Context calculations
        is_failed = is_active_failure(scan_row.get(SCAN_RESULT_COL, ""))
        status_str = "Active" if is_failed else "Inactive"
        
        already_present = is_already_in_exceptions(s_pasteable, base_exceptions_blob)
        already_present_str = "TRUE" if already_present else "FALSE"
        
        needs_ai = "TRUE" if (status_str == "Active" and already_present_str == "FALSE") else "FALSE"
        
        ledger_dict[exception_id] = {
            "Exception_ID": exception_id,
            "IP Address": scan_row.get(SCAN_HOST_COL, s_ip),
            "Host Name": s_hostname,
            "Source Sheet": source_sheet_name,
            "Plugin ID": scan_row.get(SCAN_PLUGIN_COL, ""),
            "STIG": scan_row.get(SCAN_STIG_COL, ""),
            "FINDING": s_finding if s_finding else "N/A",
            "CAT": scan_row.get('cat', ""),
            "Severity": scan_row.get('severity', ""),
            "Result": scan_row.get(SCAN_RESULT_COL, ""),
            "Short Desc": scan_row.get(SCAN_SHORT_DESC, ""),
            "Plugin Name": scan_row.get(SCAN_PLUGIN_NAME, ""),
            "Pasteable": s_pasteable,
            "Compliance Reference": scan_row.get('compliance reference', ""),
            "First Seen Date": current_run_date,
            "Last Seen Date": current_run_date,
            "Exception Status": status_str,
            "Already In Benchmark Exceptions List?": already_present_str,
            "Needs AI Draft?": needs_ai,
            "AI Packet File": "",
            "AI Response File": "",
            "AI Draft Imported?": "FALSE",
            "Reviewer Status": "Not Started",
            "Approved to Append?": "FALSE",
            "Appended to Master?": "FALSE",
            "Append Date": "",
            "Notes": ""
        }
        
    ledger_df = pd.DataFrame(list(ledger_dict.values()))

    # --------------------------------------------------------------------------
    # STEP 6: Prompt for Destination & Apply Workbook Layout Formatting
    # --------------------------------------------------------------------------
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    default_filename = f"Consolidated_STIG_Merge_Report_{timestamp}.xlsx"
    
    print("\n[*] Opening File Explorer Window: Choose destination folder to SAVE report...")
    save_dest = filedialog.asksaveasfilename(
        title="Save Consolidated STIG Report",
        initialfile=default_filename,
        defaultextension=".xlsx",
        filetypes=[("Excel Workbook", "*.xlsx")]
    )
    
    if not save_dest:
        print("[-] Save operation canceled by operator. Defaulting to local workspace directory.")
        save_dest = Path.cwd() / default_filename
    else:
        save_dest = Path(save_dest)
        
    print(f"[*] Compiling worksheets into spreadsheet: {save_dest}")
    
    with pd.ExcelWriter(save_dest, engine='openpyxl') as writer:
        sheet1_df.to_excel(writer, sheet_name="Master Tracker Evaluation", index=False)
        raw_scan_pool_df.drop(columns=['norm_ip'], errors='ignore').to_excel(writer, sheet_name="Raw Scan Data Pool", index=False)
        sheet3_df.to_excel(writer, sheet_name="STIG ID Summary", index=False)
        sheet4_df.to_excel(writer, sheet_name="Plugin ID Summary", index=False)
        # FIX: Appending Exception Ledger worksheet data frames safely
        ledger_df.to_excel(writer, sheet_name="Exception Ledger", index=False)
        
        # Format the spreadsheet cells cleanly
        for sheet_name, worksheet in writer.sheets.items():
            print(f"   [>] Optimizing column spacing and text wrapping rules on: '{sheet_name}'")
            
            if hasattr(worksheet, 'sheet'):
                worksheet = worksheet.sheet
                
            # Freeze the top header row on the current worksheet
            worksheet.freeze_panes = "A2"

            # FIX: Apply Excel Data Range Auto-Filters dynamically across all worksheets
            worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
            
            # Form long text wrap rules lists arrays matching criteria
            long_wrap_cols = ["FINDING", "Short Desc", "Plugin Name", "Pasteable", "Compliance Reference", "Notes", "Benchmark Exceptions List"]

            # Iterate through columns using safe index logic
            for col_idx in range(1, worksheet.max_column + 1):
                max_len = 0
                col_letter = get_column_letter(col_idx)
                header_value = str(worksheet.cell(row=1, column=col_idx).value).strip()
                
                for row_idx in range(1, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)

                    # Force cell alignments layout
                    if header_value in long_wrap_cols:
                        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
                    else:
                        cell.alignment = Alignment(wrap_text=False, vertical="top", horizontal="left")
                    
                    if cell.value is not None:
                        lines = str(cell.value).split('\n')
                        for line in lines:
                            if len(line) > max_len:
                                max_len = len(line)
                                
                calculated_width = max_len + 3
                # Let wrapped long description details columns breathe with wide layout spaces
                if header_value in long_wrap_cols and calculated_width > 45:
                    worksheet.column_dimensions[col_letter].width = 45
                elif calculated_width > 90:
                    worksheet.column_dimensions[col_letter].width = 90
                else:
                    worksheet.column_dimensions[col_letter].width = max(calculated_width, 11)

    print(f"[+] Process complete! Saved file context output directly to: {save_dest.name}")
    root.destroy()

# ==============================================================================
# RUNNER RUN INTERFACE
# ==============================================================================
if __name__ == "__main__": 
    merge_deviation_sheets()