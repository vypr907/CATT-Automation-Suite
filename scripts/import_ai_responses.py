#!/usr/bin/env python3
import os
import sys
import re
import json
import shutil
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment

# Import Tkinter for Graphical User Interface elements
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog

# ==============================================================================
# EXPECTED SCHEMA CONSTANTS
# ==============================================================================
SHEET_MASTER = "Master Tracker Evaluation"
SHEET_LEDGER = "Exception Ledger"

# Target Master Sheet Columns
MASTER_IP = "IP Address"
MASTER_EXCEPT_PENDING = "New Exceptions Pending AI Draft"
MASTER_AI_JUST = "AI Draft Justification Addendum"
MASTER_AI_MIT = "AI Draft Mitigating Controls Addendum"
MASTER_AI_REF = "AI Source References Used"
MASTER_AI_GAPS = "AI Assumptions / Gaps"
MASTER_AI_CONF = "AI Confidence"
MASTER_REV_STATUS = "Reviewer Status"
MASTER_READY_APP = "Ready to Append?"
MASTER_APP_COMP = "Append Completed?"

# Target Exception Ledger Columns
LEDGER_ID = "Exception_ID"
LEDGER_IP = "IP Address"
LEDGER_NEEDS_AI = "Needs AI Draft?"
LEDGER_RESP_FILE = "AI Response File"
LEDGER_IMPORTED = "AI Draft Imported?"
LEDGER_REV_STATUS = "Reviewer Status"

# ==============================================================================
# CORE CORE DATA HELPER FUNCTIONS
# ==============================================================================
def normalize_ip(ip_val) -> str:
    """Safely converts an IP address tracking cell value to a clean string."""
    if ip_val is None:
        return ""
    return str(ip_val).strip().lower()

def normalize_bool(val) -> bool:
    """Normalizes variation blocks in Excel fields to true boolean states."""
    if val is None:
        return False
    clean_str = str(val).strip().upper()
    return clean_str in ["TRUE", "1", "YES", "Y"]

def get_header_map(ws) -> dict:
    """Creates a 1-based column mapping dict from the header row of a worksheet."""
    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val is not None:
            header_map[str(val).strip()] = col_idx
    return header_map

def validate_required_columns(header_map, required_columns, sheet_name):
    """Verifies that all mandatory tracking columns exist within the header map."""
    missing = [col for col in required_columns if col not in header_map]
    if missing:
        messagebox.showerror("Schema Validation Error", f"Missing required columns in tab '{sheet_name}':\n{missing}")
        sys.exit(1)

def join_list_for_excel(value) -> str:
    """Converts arrays or strings from JSON fields into newline-separated values."""
    if isinstance(value, list):
        return "\n".join([str(item).strip() for item in value if str(item).strip()])
    if value is None:
        return ""
    return str(value).strip()

# ==============================================================================
# GRAPHICAL DIALOG CONFIGURATION BLOCK
# ==============================================================================
class ImportOptionsDialog(simpledialog.Dialog):
    """Custom Tkinter configuration modal window for pipeline ingestion rules."""
    def body(self, master):
        self.title("AI Response Importer Configurations")
        
        tk.Label(master, text="Configure Import Behaviors:", font=('Helvetica', 10, 'bold')).grid(row=0, columnspan=2, sticky="w", pady=5)
        
        # Backup Checkbox
        self.var_backup = tk.BooleanVar(value=True)
        tk.Checkbutton(master, text="Create a timestamped backup copy before modifying workbook", variable=self.var_backup).grid(row=1, columnspan=2, sticky="w", pady=2)
        
        # Overwrite Checkbox
        self.var_overwrite = tk.BooleanVar(value=False)
        tk.Checkbutton(master, text="Overwrite existing data in AI draft review columns", variable=self.var_overwrite).grid(row=2, columnspan=2, sticky="w", pady=2)
        
        # Dry Run Checkbox
        self.var_dry_run = tk.BooleanVar(value=False)
        tk.Checkbutton(master, text="Dry Run Mode (Simulate and log outputs without modifying files)", variable=self.var_dry_run).grid(row=3, columnspan=2, sticky="w", pady=2)
        
        # Execution Cap Limitation Input
        tk.Label(master, text="Limitation Cap (Max response files to process, 0 for ALL):").grid(row=4, column=0, sticky="w", pady=5)
        self.entry_limit = tk.Entry(master, width=10)
        self.entry_limit.insert(0, "0")
        self.entry_limit.grid(row=4, column=1, sticky="w", pady=5)
        
        return self.entry_limit

    def apply(self):
        self.backup = self.var_backup.get()
        self.overwrite = self.var_overwrite.get()
        self.dry_run = self.var_dry_run.get()
        
        limit_val = self.entry_limit.get().strip()
        try:
            self.limit = int(limit_val) if (limit_val and limit_val != "0") else None
        except ValueError:
            self.limit = None

# ==============================================================================
# UNPARSED RESPONSE PARSING MATRIX
# ==============================================================================
def extract_json_from_file(file_path: Path) -> dict:
    """Reads raw source logs pulling out text structures wrapped inside Markdown block tags."""
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        content = f.read().strip()
        
    # Attempt to locate standard backtick markdown fence strings block
    json_match = re.search(r'```json\s*(\{.*?\})\s*```', content, re.DOTALL | re.IGNORECASE)
    if json_match:
        content = json_match.group(1)
    else:
        # Fallback to standard trailing fence patterns
        general_code_match = re.search(r'```\s*(\{.*?\})\s*```', content, re.DOTALL)
        if general_code_match:
            content = general_code_match.group(1)
            
    # Try parsing clean raw string block
    try:
        return json.loads(content.strip())
    except json.JSONDecodeError:
        # Brute force search look for structural opening curly bracket parameters
        start_idx = content.find('{')
        end_idx = content.rfind('}')
        if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
            try:
                return json.loads(content[start_idx:end_idx+1])
            except json.JSONDecodeError:
                pass
        raise ValueError("File content could not be cleanly serialized into a valid JSON object map structure.")

def validate_response_payload(payload: dict, file_name: str) -> str:
    """Validates structural constraints of incoming JSON payloads before indexing."""
    if not isinstance(payload, dict):
        return "Payload is not an object map grid dictionary configuration."
        
    if "ip_address" not in payload or not str(payload.get("ip_address")).strip():
        return "Mandatory parameter key 'ip_address' missing or structural string empty."
        
    if "exceptions_reviewed" not in payload or not isinstance(payload.get("exceptions_reviewed"), list):
        return "Mandatory array index property 'exceptions_reviewed' missing or malformed tracking matrix."
        
    if "draft_justification_addendum" not in payload or "draft_mitigating_controls_addendum" not in payload:
        return "Missing core required field 'draft_justification_addendum' or 'draft_mitigating_controls_addendum'."
        
    conf = str(payload.get("confidence", "")).strip().upper()
    if conf not in ["HIGH", "MEDIUM", "LOW"]:
        return f"Provided validation confidence configuration string metric '{conf}' is invalid."
        
    return ""

# ==============================================================================
# MAIN SYSTEM INGESTION CORE ENGINE
# ==============================================================================
def main():
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    
    # 1. Capture Data Sources Graphically
    print("[*] Launching Workbook File Ingestion Dialog Selector Window...")
    wb_selected = filedialog.askopenfilename(
        title="Select Consolidated STIG Workbook Target (.xlsx)",
        filetypes=[("Excel Workbooks", "*.xlsx")]
    )
    if not wb_selected:
        print("[-] Execution stopped: Target workbook placeholder input file missing.")
        sys.exit(0)
    workbook_path = Path(wb_selected)
    
    print("[*] Launching Directory Target Selector for Gemini Response Files...")
    dir_selected = filedialog.askdirectory(title="Select Folder Containing Gemini Response Documents")
    if not dir_selected:
        print("[-] Execution stopped: Target responses root workspace directory not supplied.")
        sys.exit(0)
    responses_dir = Path(dir_selected)
    
    # 2. Launch Graphical Configurations Modal Context Window Form
    dialog_form = ImportOptionsDialog(root)
    opt_backup = getattr(dialog_form, 'backup', True)
    opt_overwrite = getattr(dialog_form, 'overwrite', False)
    opt_dry_run = getattr(dialog_form, 'dry_run', False)
    opt_limit = getattr(dialog_form, 'limit', None)
    
    # Run dashboard counter performance analytics metrics variables tracking parameters
    metrics = {"found": 0, "processed": 0, "skipped": 0, "master_updated": 0, "ledger_updated": 0, "warnings": []}
    
    print("[*] Accessing target file storage platform engine...")
    try:
        wb = openpyxl.load_workbook(workbook_path)
    except Exception as e:
        messagebox.showerror("File Lock Error", f"Unable to safely map connection to workbook target spreadsheet file:\n{e}")
        sys.exit(1)
        
    if SHEET_MASTER not in wb.sheetnames or SHEET_LEDGER not in wb.sheetnames:
        messagebox.showerror("Workbook Structural Error", f"Workbook configuration does not contain required worksheets:\n- '{SHEET_MASTER}'\n- '{SHEET_LEDGER}'")
        sys.exit(1)
        
    ws_master = wb[SHEET_MASTER]
    ws_ledger = wb[SHEET_LEDGER]
    
    # Parse mappings indexes
    master_map = get_header_map(ws_master)
    ledger_map = get_header_map(ws_ledger)
    
    # Validate column integrity constraints across sheets baseline arrays parameters
    validate_required_columns(master_map, [MASTER_IP, MASTER_EXCEPT_PENDING, MASTER_AI_JUST, MASTER_AI_MIT, MASTER_AI_REF, MASTER_AI_GAPS, MASTER_AI_CONF, MASTER_REV_STATUS, MASTER_READY_APP, MASTER_APP_COMP], SHEET_MASTER)
    validate_required_columns(ledger_map, [LEDGER_ID, LEDGER_IP, LEDGER_NEEDS_AI, LEDGER_RESP_FILE, LEDGER_IMPORTED, LEDGER_REV_STATUS], SHEET_LEDGER)
    
    # Scan target directory patterns for valid types
    valid_extensions = {".json", ".txt", ".md"}
    response_files = [p for p in responses_dir.iterdir() if p.is_file() and p.suffix.lower() in valid_extensions]
    metrics["found"] = len(response_files)
    
    if not response_files:
        messagebox.showinfo("Execution Terminal Alert", f"Zero applicable response log files (.json, .txt, .md) discovered in directory:\n'{responses_dir.name}'")
        sys.exit(0)
        
    print(f"[+] Discovered {len(response_files)} files to evaluate for workflow import sequence rules.")
    
    # Alignment styling template configuration block helper instantiation
    wrap_align = Alignment(wrap_text=True, vertical="top", horizontal="left")
    
    # --------------------------------------------------------------------------
    # PROCESSING REVOLUTION MATRIX LOOP
    # --------------------------------------------------------------------------
    for r_path in response_files:
        if opt_limit is not None and metrics["processed"] >= opt_limit:
            print(f"[*] Limiting scope boundaries: Hit processing run threshold constraint index cap ({opt_limit})")
            break
            
        print(f"   [>] Evaluating source log context artifact -> '{r_path.name}'")
        try:
            payload = extract_json_from_file(r_path)
            err_msg = validate_response_payload(payload, r_path.name)
            if err_msg:
                raise ValueError(err_msg)
        except Exception as e:
            w_msg = f"Skipped parsing failure inside '{r_path.name}': {e}"
            print(f"   [-] Warning: {w_msg}")
            metrics["warnings"].append(w_msg)
            metrics["skipped"] += 1
            continue
            
        # Target matching variables parsing execution
        target_ip = normalize_ip(payload.get("ip_address"))
        reviewed_exceptions_list = payload.get("exceptions_reviewed", [])
        
        # Locate matching row inside Sheet 1 summary dataset
        master_row_idx = None
        for r_idx in range(2, ws_master.max_row + 1):
            grid_ip = normalize_ip(ws_master.cell(row=r_idx, column=master_map[MASTER_IP]).value)
            if grid_ip == target_ip:
                master_row_idx = r_idx
                break
                
        if not master_row_idx:
            w_msg = f"Asset target IP key '{payload.get('ip_address')}' referenced in '{r_path.name}' does not align with any record inside '{SHEET_MASTER}'."
            print(f"   [-] Warning: {w_msg}")
            metrics["warnings"].append(w_msg)
            metrics["skipped"] += 1
            continue
            
        # Check overwrite block validation locks
        existing_val = ws_master.cell(row=master_row_idx, column=master_map[MASTER_AI_JUST]).value
        if existing_val and not opt_overwrite:
            print(f"   [>] Skipping import routine on target asset row {master_row_idx}: Content already exists and overwrite flag is disabled.")
            metrics["skipped"] += 1
            continue
            
        # --------------------------------------------------------------------------
        # COMMIT UPDATE OPERATIONS TO IN-MEMORY OBJECT MODELS
        # --------------------------------------------------------------------------
        metrics["processed"] += 1
        
        # Write asset metrics telemetry mapping data properties safely down to grid row
        if not opt_dry_run:
            ws_master.cell(row=master_row_idx, column=master_map[MASTER_AI_JUST], value=str(payload.get("draft_justification_addendum")).strip()).alignment = wrap_align
            ws_master.cell(row=master_row_idx, column=master_map[MASTER_AI_MIT], value=str(payload.get("draft_mitigating_controls_addendum")).strip()).alignment = wrap_align
            ws_master.cell(row=master_row_idx, column=master_map[MASTER_AI_REF], value=join_list_for_excel(payload.get("source_references_used"))).alignment = wrap_align
            ws_master.cell(row=master_row_idx, column=master_map[MASTER_AI_GAPS], value=join_list_for_excel(payload.get("assumptions_or_gaps"))).alignment = wrap_align
            ws_master.cell(row=master_row_idx, column=master_map[MASTER_AI_CONF], value=str(payload.get("confidence")).strip()).alignment = wrap_align
            ws_master.cell(row=master_row_idx, column=master_map[MASTER_REV_STATUS], value="Pending Review").alignment = wrap_align
            
            # Explicit Safety Constraint: Keep core operational variables un-modified
            ws_master.cell(row=master_row_idx, column=master_map[MASTER_READY_APP], value="FALSE")
            
        metrics["master_updated"] += 1
        
        # Process individual Exception Ledger row validations tracking links updates
        for ex_item in reviewed_exceptions_list:
            ex_id = str(ex_item.get("exception_id", "")).strip()
            if not ex_id:
                continue
                
            ledger_row_idx = None
            for r_idx in range(2, ws_ledger.max_row + 1):
                grid_ex_id = str(ws_ledger.cell(row=r_idx, column=ledger_map[LEDGER_ID]).value).strip()
                if grid_ex_id == ex_id:
                    ledger_row_idx = r_idx
                    break
                    
            if not ledger_row_idx:
                w_msg = f"Vulnerability Exception_ID reference context '{ex_id}' listed inside '{r_path.name}' cannot be mapped back to any entry in '{SHEET_LEDGER}'."
                print(f"   [-] Warning: {w_msg}")
                metrics["warnings"].append(w_msg)
                continue
                
            if not opt_dry_run:
                ws_ledger.cell(row=ledger_row_idx, column=ledger_map[LEDGER_RESP_FILE], value=str(r_path.name)).alignment = wrap_align
                ws_ledger.cell(row=ledger_row_idx, column=ledger_map[LEDGER_IMPORTED], value="TRUE").alignment = wrap_align
                ws_ledger.cell(row=ledger_row_idx, column=ledger_map[LEDGER_REV_STATUS], value="Pending Review").alignment = wrap_align
                
            metrics["ledger_updated"] += 1

    # --------------------------------------------------------------------------
    # POST-PROCESSING RE-EVALUATION PASSTHROUGH (Calculate Remaining Gaps)
    # --------------------------------------------------------------------------
    if not opt_dry_run:
        print("[*] Sweeping ledger grids to recalculate open tracking anomalies status rules...")
        for r_idx in range(2, ws_master.max_row + 1):
            m_ip = normalize_ip(ws_master.cell(row=r_idx, column=master_map[MASTER_IP]).value)
            if not m_ip:
                continue
                
            # Scan ledger rows to calculate if any outstanding items remain un-resolved
            unimported_gaps_exist = False
            for l_idx in range(2, ws_ledger.max_row + 1):
                l_ip = normalize_ip(ws_ledger.cell(row=l_idx, column=ledger_map[LEDGER_IP]).value)
                if l_ip == m_ip:
                    needs_ai = normalize_bool(ws_ledger.cell(row=l_idx, column=ledger_map[LEDGER_NEEDS_AI]).value)
                    was_imported = normalize_bool(ws_ledger.cell(row=l_idx, column=ledger_map[LEDGER_IMPORTED]).value)
                    
                    if needs_ai and not was_imported:
                        unimported_gaps_exist = True
                        break
                        
            # Update the macro workflow tracking flag column based on current findings status
            ws_master.cell(row=r_idx, column=master_map[MASTER_EXCEPT_PENDING], value="TRUE" if unimported_gaps_exist else "FALSE")

    # --------------------------------------------------------------------------
    # BACKUP CREATION AND STORAGE ENFORCEMENT DISK WRITE LOCKS
    # --------------------------------------------------------------------------
    print("\n[*] Summary calculations complete. Terminating ledger iteration matrix.")
    if opt_dry_run:
        summary_txt = f"Dry Run Complete!\n\n- Responses Analyzed: {metrics['processed']}\n- Master Rows To Adjust: {metrics['master_updated']}\n- Ledger Items To Tag: {metrics['ledger_updated']}\n- Files Skipped: {metrics['skipped']}\n\nZero modifications were written back to your Excel spreadsheet data grid storage."
        messagebox.showinfo("Simulation Success Dashboard", summary_txt)
        sys.exit(0)
        
    if metrics["processed"] == 0:
        messagebox.showinfo("Import Operation Complete", "Zero records qualified for modification data ingestion routines.\nSpreadsheet matches left unchanged.")
        sys.exit(0)
        
    if opt_backup:
        t_stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"{workbook_path.stem}_backup_{t_stamp}{workbook_path.suffix}"
        backup_path = workbook_path.parent / backup_name
        print(f"[*] Saving archival timestamped backup configuration target to -> '{backup_name}'")
        try:
            shutil.copy2(workbook_path, backup_path)
        except Exception as e:
            messagebox.showwarning("Backup Exception Error", f"Archival backup creation skipped because of target access permissions configuration failures:\n{e}")
            
    print(f"[*] Invoking Openpyxl engine layout save operation down to disk file coordinates -> '{workbook_path.name}'")
    try:
        wb.save(workbook_path)
        success_txt = f"Import Process Executed Successfully!\n\n- Response Artifacts Found: {metrics['found']}\n- Valid Files Digested: {metrics['processed']}\n- Input Files Skipped/Bypassed: {metrics['skipped']}\n- Master Assets Summaries Updated: {metrics['master_updated']}\n- Ledger Entry Reference Cells Linked: {metrics['ledger_updated']}\n- Warnings Generated: {len(metrics['warnings'])}"
        messagebox.showinfo("Success Dashboard", success_txt)
    except Exception as e:
        messagebox.showerror("Spreadsheet Write Access Lock Error", f"Critical Storage Failure: Excel file save lock blocked.\nClose the spreadsheet immediately if open in Excel, then re-execute script:\n{e}")
        sys.exit(1)
        
    root.destroy()

if __name__ == "__main__":
    main()